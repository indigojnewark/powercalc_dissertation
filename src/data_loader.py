"""data_loader.py

Loads and inspects LFQ intensity data from MaxQuant Excel output.
Handles multi-sheet workbooks and extracts valid protein intensity
values for pairwise group comparisons.
"""

import openpyxl


def inspect_workbook(filepath: str, max_preview_rows: int = 3) -> None:
    """
    Print sheet names, dimensions, and a row preview for a given Excel file.

    Parameters
    ----------
    filepath : str
        Path to the Excel workbook.
    max_preview_rows : int
        Number of rows to preview per sheet (default: 3).
    """
    with openpyxl.load_workbook(filepath, read_only=True) as wb:
        if not wb.sheetnames:
            print(f"{filepath}: no sheets found.")
            return
        print(f"\n=== {filepath} ===")
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"  Sheet: {sheet_name}, dims: {ws.dimensions}")
            rows = list(ws.iter_rows(max_row=max_preview_rows, values_only=True))
            for row in rows:
                print(f"    {row[:12]}")


def load_sheet(filepath: str, sheet_name: str) -> list:
    """
    Load all rows from a named sheet as a list of tuples.

    Parameters
    ----------
    filepath : str
        Path to the Excel workbook.
    sheet_name : str
        Name of the sheet to load.

    Returns
    -------
    list of tuple
        All rows including the header row.

    Raises
    ------
    ValueError
        If the requested sheet is not found in the workbook.
    """
    with openpyxl.load_workbook(filepath, read_only=True) as wb:
        if sheet_name not in wb.sheetnames:
            raise ValueError(
                f"Sheet '{sheet_name}' not found in {filepath}. "
                f"Available sheets: {wb.sheetnames}"
            )
        ws = wb[sheet_name]
        return list(ws.iter_rows(values_only=True))


def get_lfq_values(
    rows: list,
    group_a_cols: list,
    group_b_cols: list,
    min_valid: int = 2,
) -> list:
    """
    Extract LFQ intensities for two groups from row data.

    Skips the header row (index 0) and filters proteins to those
    with at least `min_valid` non-zero, non-null values in both groups.

    Parameters
    ----------
    rows : list of tuple
        All rows from the sheet, including the header at index 0.
    group_a_cols : list of int
        Column indices for group A LFQ intensities.
    group_b_cols : list of int
        Column indices for group B LFQ intensities.
    min_valid : int
        Minimum number of valid (non-zero) values required per group.

    Returns
    -------
    list of tuple
        Each element is (group_a_vals, group_b_vals) for a qualifying protein.
    """
    data = []
    for row in rows[1:]:  # rows[0] is the header
        a_vals = [row[i] for i in group_a_cols if row[i] and row[i] > 0]
        b_vals = [row[i] for i in group_b_cols if row[i] and row[i] > 0]
        if len(a_vals) >= min_valid and len(b_vals) >= min_valid:
            data.append((a_vals, b_vals))
    return data
